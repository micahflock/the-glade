#!/usr/bin/env python3
"""One-shot generator for encounters/winter_farm_encounter.xlsx"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/the-glade/encounters/winter_farm_encounter.xlsx"

# ── palette ────────────────────────────────────────────────────────────────
TRACKER_HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SHADE_FILL = PatternFill("solid", fgColor="F2F2F2")

BOLD = Font(name="Arial", size=10, bold=True)
BOLD_WHITE = Font(name="Arial", size=12, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True)
DATA = Font(name="Arial", size=10)

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

COLS = 8  # A..H

# ── data ───────────────────────────────────────────────────────────────────
# Rolled by scripts/roll_dice.py
rolled_npcs = [
    {"name": "Pheromone Crowner", "ac": 15, "hp": 134, "speed": "30/c30", "init": 8,  "notes": ""},
    {"name": "Leafcutter Worker 1 [A]", "ac": 13, "hp": 60, "speed": "30/c30", "init": 7, "notes": "Group A"},
    {"name": "Leafcutter Worker 2 [A]", "ac": 13, "hp": 69, "speed": "30/c30", "init": 7, "notes": "Group A"},
    {"name": "Leafcutter Worker 3 [A]", "ac": 13, "hp": 49, "speed": "30/c30", "init": 7, "notes": "Group A"},
    {"name": "Leafcutter Worker 4 [A]", "ac": 13, "hp": 62, "speed": "30/c30", "init": 7, "notes": "Group A"},
    {"name": "Skull-Cracker",     "ac": 16, "hp": 162, "speed": "30/c20", "init": 4, "notes": ""},
]

pcs = [
    {"name": "Tabatha Starr (Heather)",   "ac": 14, "hp_max": 31, "speed": "30", "init_mod": 5, "notes": "Dread Ambusher rd 1"},
    {"name": "Jiminy 'Drix' Hendrix (Katie)", "ac": 15, "hp_max": 21, "speed": "30", "init_mod": 3, "notes": ""},
    {"name": "Nancy Alfons (ROWYN)",      "ac": 15, "hp_max": 22, "speed": "40", "init_mod": 3, "notes": "Mobile: +10 if Dash"},
]

# ── workbook ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Winter Farm Encounter"

# Column widths
widths = {"A": 8, "B": 28, "C": 6, "D": 9, "E": 11, "F": 8, "G": 22, "H": 28}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

row = 1

def set_row(r, values, fills=None, font=DATA, align=WRAP, border=BORDER, height=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = font
        c.alignment = align
        c.border = border
        if fills:
            c.fill = fills
    if height:
        ws.row_dimensions[r].height = height

def merge_row(r, label, content, label_font=BOLD, content_font=DATA, fill=None, height=None):
    """Label in col B, merged content in C–H. Col A left blank."""
    for col in range(1, COLS + 1):
        c = ws.cell(row=r, column=col)
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.cell(row=r, column=2, value=label).font = label_font
    ws.cell(row=r, column=2).alignment = WRAP
    ws.cell(row=r, column=3, value=content).font = content_font
    ws.cell(row=r, column=3).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=COLS)
    if height:
        ws.row_dimensions[r].height = height

def section_header(r, text):
    ws.cell(row=r, column=1, value=text)
    for col in range(1, COLS + 1):
        c = ws.cell(row=r, column=col)
        c.fill = SECTION_HEADER_FILL
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    ws.row_dimensions[r].height = 22

def subheader(r, text):
    ws.cell(row=r, column=1, value=text)
    for col in range(1, COLS + 1):
        c = ws.cell(row=r, column=col)
        c.fill = TRACKER_HEADER_FILL
        c.font = BOLD
        c.alignment = LEFT
        c.border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    ws.row_dimensions[r].height = 18

# ── SECTION 1: INITIATIVE TRACKER ──────────────────────────────────────────
headers = ["Init", "Name", "AC", "HP Max", "HP Current", "Speed", "Status/Conditions", "Notes"]
set_row(row, headers, fills=TRACKER_HEADER_FILL, font=BOLD, align=CENTER, height=20)
row += 1

# Sort: NPCs by init desc, then PCs (blank init last)
rolled_npcs_sorted = sorted(rolled_npcs, key=lambda n: -n["init"])

shade = False
for n in rolled_npcs_sorted:
    values = [n["init"], n["name"], n["ac"], n["hp"], n["hp"], n["speed"], "", n["notes"]]
    fill = SHADE_FILL if shade else None
    set_row(row, values, fills=fill, height=18)
    shade = not shade
    row += 1

for p in pcs:
    values = ["", p["name"], p["ac"], p["hp_max"], "", p["speed"], "", f"Init {p['init_mod']:+d}  {p['notes']}".strip()]
    fill = SHADE_FILL if shade else None
    set_row(row, values, fills=fill, height=18)
    shade = not shade
    row += 1

# Freeze panes after tracker so it stays pinned.
ws.freeze_panes = ws.cell(row=row, column=1)

# ── SECTION 2: NPC REFERENCE ───────────────────────────────────────────────
row += 2
section_header(row, "NPC REFERENCE"); row += 1

def write_creature(start_row, creature):
    r = start_row
    merge_row(r, "Name", creature["title"], height=18); r += 1
    merge_row(r, "Stats", creature["stats"], height=32); r += 1
    if creature.get("defenses"):
        merge_row(r, "Defenses", creature["defenses"], height=32); r += 1
    for section_label, entries in creature["entries"]:
        if section_label:
            subheader(r, section_label); r += 1
        for name, text in entries:
            est_lines = max(1, len(text) // 70 + text.count("\n") + 1)
            merge_row(r, name, text, height=max(18, est_lines * 15)); r += 1
    return r + 1  # blank row after

crowner = {
    "title": "Pheromone Crowner (Spring) — Medium Monstrosity (ant), Lawful Evil",
    "stats": ("AC 15 (natural armor) | HP 134 (17d8+51) | Speed 30 ft., climb 30 ft. | "
              "STR 14 (+2) DEX 14 (+2) CON 16 (+3) INT 10 (+0) WIS 16 (+3) CHA 16 (+3)"),
    "defenses": ("Skills: Perception +6, Insight +6 | "
                 "Senses: darkvision 60 ft., tremorsense 30 ft., passive Perception 13 | "
                 "Languages: Ant-cant (chemical), Common (halting) | "
                 "CR 5 (1800 XP) | PB +3"),
    "entries": [
        ("TRAITS", [
            ("Heated Body",
             "A creature that touches the Crowner or hits her with a melee attack while within 5 ft. "
             "takes 2d6 fire damage. Sheds bright light 5 ft. and dim light another 5 ft."),
            ("Pheromone Sense",
             "Aware of the location and emotional state of every fire ant within 120 ft. "
             "Senses emotional states of non-ants within 60 ft. (adv on Insight vs them)."),
            ("Visible Scent-Trails",
             "Pheromone commands glow faintly in the air for 1 round after release. "
             "Narrative only except as specified in Scorch-Trail."),
        ]),
        ("ACTIONS", [
            ("Multiattack", "The Crowner makes two Bite attacks."),
            ("Bite",
             "Melee Weapon Attack: +6 to hit, reach 5 ft., one target. "
             "Hit: 11 (2d8+2) piercing plus 3 (1d6) fire damage."),
            ("Scorch-Trail (Recharge 5–6)",
             "Draws a 30-ft. line, 5 ft. wide, from her. Each creature in or entering the line before "
             "her next turn: DC 13 CON save, 7 (2d6) fire on fail, half on success. On fail, also marked — "
             "next attack against it before the end of the Crowner's next turn has advantage. "
             "Allied ants moving along the line: +10 ft. speed for that movement, +1d6 fire on next melee "
             "hit before end of turn."),
        ]),
        ("BONUS ACTIONS", [
            ("Crown-Scent (1/turn)",
             "One ally she can see within 30 ft. gains advantage on its next attack roll before end of "
             "its next turn, OR may immediately move up to half its speed without provoking OAs."),
        ]),
    ],
}

skullcracker = {
    "title": "Skull-Cracker (Spring) — Large Monstrosity (ant), Lawful Evil",
    "stats": ("AC 16 (natural armor) | HP 162 (16d10+64) | Speed 30 ft., climb 20 ft. | "
              "STR 20 (+5) DEX 10 (+0) CON 18 (+4) INT 6 (-2) WIS 12 (+1) CHA 10 (+0)"),
    "defenses": ("Skills: Athletics +8 | "
                 "Senses: darkvision 60 ft., tremorsense 30 ft., passive Perception 11 | "
                 "Languages: Ant-cant (chemical, simple) | "
                 "CR 6 (2300 XP) | PB +3"),
    "entries": [
        ("TRAITS", [
            ("Heated Body",
             "A creature that touches the Skull-Cracker or hits him with a melee attack while within "
             "5 ft. takes 2d6 fire damage. Mandibles glow forge-hot; dim light to 10 ft."),
            ("Head-Heavy",
             "Disadvantage on DEX saves vs prone. Advantage on STR saves/checks to resist being moved, "
             "grappled, or shoved."),
            ("Forge-Mandibles",
             "Bite ignores resistance to piercing damage. Non-magical armor broken by Crusher's Bite is "
             "charred — requires a proper forge to restore."),
        ]),
        ("ACTIONS", [
            ("Multiattack",
             "Two Crusher's Bite attacks, or one Crusher's Bite + one Mandible Slam. May replace one "
             "Crusher's Bite with Sear-Snap if available."),
            ("Crusher's Bite",
             "Melee Weapon Attack: +8 to hit, reach 10 ft., one target. "
             "Hit: 14 (2d8+5) piercing plus 7 (2d6) fire. On hit vs non-magical armor, armor cracks: "
             "AC -1 until repaired (stacks to -3 per creature)."),
            ("Mandible Slam",
             "Melee Weapon Attack: +8 to hit, reach 10 ft., one target. "
             "Hit: 12 (2d6+5) bludgeoning. Target DC 15 STR save or grappled (escape DC 15). "
             "Only one target grappled this way at a time."),
            ("Sear-Snap (Recharge 5–6)",
             "Melee Weapon Attack: +8 to hit, reach 10 ft., one target. "
             "Hit: 16 (3d8+3) piercing plus 7 (2d6) fire. If target is prone or grappled, fire damage "
             "doubles (to 14 / 4d6) and wound smolders: target takes 3 (1d6) fire at start of each of "
             "its turns until it succeeds on a DC 13 CON save at the end of its turn."),
        ]),
    ],
}

leafcutter = {
    "title": "Leafcutter Worker — Medium Monstrosity (ant), Neutral",
    "stats": ("AC 13 (natural armor) | HP 58 (9d8+18) | Speed 30 ft., climb 30 ft. | "
              "STR 14 (+2) DEX 12 (+1) CON 14 (+2) INT 6 (-2) WIS 10 (+0) CHA 8 (-1)"),
    "defenses": ("Senses: darkvision 60 ft., tremorsense 30 ft., passive Perception 10 | "
                 "Languages: Ant-cant (chemical) | CR 1/2 (100 XP) | PB +2"),
    "entries": [
        ("TRAITS", [
            ("Beast of Burden",
             "Carrying capacity as if one size larger. Can drag/haul up to 5x normal capacity without "
             "penalty."),
            ("Leaf-Slicer Mandibles",
             "Outside combat, mandibles cleanly cut leaves, cloth, paper, thin wood. Can strip a leaf "
             "in seconds."),
            ("Reluctant Soldier",
             "When every fire ant commander/overseer in the encounter is dead or fled, at the start of "
             "its next turn the worker makes a DC 10 WIS save. On fail: stands down (flees, surrenders, "
             "or parlays). On success: fights on with disadvantage on attack rolls until a new commander "
             "arrives or combat ends."),
        ]),
        ("ACTIONS", [
            ("Mandible Snip",
             "Melee Weapon Attack: +4 to hit, reach 5 ft., one target. "
             "Hit: 6 (1d8+2) slashing damage."),
        ]),
    ],
}

row = write_creature(row, crowner)
row = write_creature(row, skullcracker)
row = write_creature(row, leafcutter)

# ── SECTION 3: PC REFERENCE ────────────────────────────────────────────────
row += 1
section_header(row, "PC REFERENCE"); row += 1

def write_pc(start_row, pc):
    r = start_row
    merge_row(r, "Name", pc["header"], height=18); r += 1
    merge_row(r, "Abilities", pc["abilities"], height=18); r += 1
    merge_row(r, "Saves", pc["saves"], height=18); r += 1
    merge_row(r, "Attacks", pc["attacks"], height=40); r += 1
    if pc.get("spells"):
        merge_row(r, "Spells", pc["spells"], height=40); r += 1
    merge_row(r, "Features", pc["features"], height=55); r += 1
    return r + 1

drix = {
    "header": "Jiminy 'Drix' Hendrix (Katie) — Bard 3 (College of Lore) | Cricket | "
              "AC 15 | HP 21 | Speed 30 | Passive Perception 12 | Init +3",
    "abilities": "STR 6 (-2) DEX 16 (+3) CON 13 (+1) INT 10 (+0) WIS 10 (+0) CHA 18 (+4)",
    "saves": "DEX +5 (prof), CHA +6 (prof). Others use ability mod.",
    "attacks": "Rapier +5, 1d8+3 piercing  |  Shortsword +5, 1d6+3 piercing  |  "
               "Spell attack +6, Spell save DC 14.",
    "spells": "Cantrips: Vicious Mockery, Prestidigitation. "
              "1st (4 slots): Healing Word, Disguise Self, Silvery Barbs, Dissonant Whispers. "
              "2nd (2 slots): Hold Person, Locate Animals or Plants.",
    "features": "Bardic Inspiration (d6, 4 uses, bonus action, 60 ft.). Jack of All Trades (+1 to "
                "non-prof checks). Cutting Words (reaction: spend Inspiration die to subtract from "
                "enemy attack/check/damage). Song Distraction (homebrew: Performance vs Insight, "
                "grabs attention; target has disadv on Perception/Investigation while you perform).",
}

nancy = {
    "header": "Nancy Alfons (ROWYN) — Rogue 3 (Arcane Trickster) | Jumping Spider | "
              "AC 15 | HP 22 | Speed 40 | Passive Perception 11 | Init +3",
    "abilities": "STR 6 (-2) DEX 16 (+3) CON 12 (+1) INT 16 (+3) WIS 12 (+1) CHA 12 (+1)",
    "saves": "DEX +5 (prof), INT +5 (prof). Others use ability mod.",
    "attacks": "Rapier +5, 1d8 piercing  |  Shortbow +5, 1d6 piercing (80/320)  |  "
               "Spell attack +5, Spell save DC 13.  |  Sneak Attack 2d6 once per turn (conditions apply).",
    "spells": "Cantrips: Mage Hand, Message, Minor Illusion. "
              "1st (2 slots): Disguise Self, Detect Magic (ritual), Identify (ritual).",
    "features": "Mobile (feat): +10 speed; Dash ignores difficult terrain; melee target can't OA you "
                "rest of turn. Fey Touched (feat). Cunning Action (bonus: Dash/Disengage/Hide). "
                "Expertise: Sleight of Hand +7, Investigation +7.",
}

tabby = {
    "header": "Tabatha Starr (Heather) — Ranger 3 (Gloomstalker) | Lone Star Tick | "
              "AC 14 | HP 31 | Speed 30 | Passive Perception 14 | Init +5 (WIS via Dread Ambusher)",
    "abilities": "STR 14 (+2) DEX 16 (+3) CON 12 (+1) INT 10 (+0) WIS 14 (+2) CHA 10 (+0)",
    "saves": "STR +4 (prof), DEX +5 (prof). Others use ability mod.",
    "attacks": "Longbow +7, 1d8+3 piercing (150/600) [Archery]  |  "
               "Dagger (thrown) +7, 1d4+3 piercing (20/60) [Archery per DM ruling]  |  "
               "Dagger (melee) +5, 1d4+3 piercing.  |  1st-level slots: 3.",
    "spells": "Spell list not yet provided by player. Ranger 3 has spells — ask Heather at the table.",
    "features": "Favored Enemy: Monstrosities (adv on WIS to track, INT to recall). Favored Terrain: "
                "Forest. Archery Fighting Style (+2 to ranged attack rolls). Dread Ambusher: +10 speed "
                "round 1; bonus attack on turn 1, +1d8 on hit; WIS added to init. Umbral Sight: "
                "darkvision 60 ft.; invisible to darkvision creatures in darkness. Sharpshooter "
                "(homebrew feat): ignore long-range disadv, ignore half/three-quarters cover, "
                "-5 to hit for +10 damage.",
}

row = write_pc(row, drix)
row = write_pc(row, nancy)
row = write_pc(row, tabby)

# ── SECTION 4: ENCOUNTER DETAILS ───────────────────────────────────────────
row += 1
section_header(row, "ENCOUNTER DETAILS"); row += 1

# Encounter title row
ws.cell(row=row, column=1, value="Winter Farm Encounter")
for col in range(1, COLS + 1):
    c = ws.cell(row=row, column=col)
    c.font = TITLE
    c.alignment = LEFT
    c.border = BORDER
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
ws.row_dimensions[row].height = 22
row += 1

notes = [
    ("Setup",
     "Whole party present: Drix (Bard), Nancy (Rogue), Tabby (Ranger). "
     "Traveling toward the Winter Farm with the Leafcutter Rebel and Leafcutter Loyalist in tow "
     "(see NPC Registry). Loyalist is planning to betray the party at earliest opportunity — "
     "GM's call whether this encounter is that betrayal payoff. Terrain not specified."),
    ("Difficulty note",
     "Combined CR 13 (Crowner 5 + Skull-Cracker 6 + 4x Worker 0.5 = CR 13) against three level-3 PCs. "
     "This is a Deadly-tier encounter. Consider positioning, cover, retreat paths."),
    ("Reluctant Soldier trigger",
     "If BOTH the Pheromone Crowner and the Skull-Cracker are killed or flee, each surviving "
     "Leafcutter Worker makes a DC 10 WIS save at the start of its next turn. "
     "Fail: stands down (flees, surrenders, parlays). Success: fights on with DISADVANTAGE on "
     "attack rolls. This is the party's lever — kill the commanders, fold the workers."),
    ("Recharge tracking",
     "Crowner: Scorch-Trail (5–6). Skull-Cracker: Sear-Snap (5–6). Roll at start of each of their turns."),
    ("Crowner tactics",
     "Opens with Scorch-Trail to carve the field and set up allies. Uses Crown-Scent every round. "
     "Positions so Heated Body aura catches flankers. Retreats below half HP (67) to rally a second wave."),
    ("Skull-Cracker tactics",
     "Opens with Mandible Slam on nearest heavy-armor target (Tabby is best candidate). Bites grappled "
     "targets. Holds Sear-Snap for the turn after a successful grapple — grappled target takes doubled "
     "fire + smolder effect. Drops grapple only for a more dangerous target."),
    ("Leafcutter tactics",
     "Shared initiative (7). Fight where directed — point them at whoever the Crowner marks. "
     "Without commander direction, hesitate and target the most obvious threat (not the most "
     "dangerous). Do NOT pursue fleeing foes."),
    ("NPC Registry cross-ref",
     "Leafcutter Rebel (disposition: Allied) and Leafcutter Loyalist (disposition: Hostile, covert) "
     "are with the party. The Rebel may aid the party; the Loyalist will seize the chance to bolt or "
     "betray."),
]

for label, text in notes:
    est_lines = max(2, len(text) // 70 + 1)
    merge_row(row, label, text, height=max(30, est_lines * 15)); row += 1

wb.save(OUT)
print(f"Wrote {OUT}")
